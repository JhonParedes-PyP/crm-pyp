import os

view_code = """
import io
import openpyxl
import datetime
import zipfile
from openpyxl.styles import Font, Alignment, Border, Side
from django.db.models import Q
from django.http import HttpResponse

@login_required
def descargar_reporte_excel(request):
    cartera = request.GET.get('cartera')
    agencia = request.GET.get('agencia')
    distritos = request.GET.getlist('distrito')
    estado_negociacion = request.GET.get('estado_negociacion')
    
    # Filtrar clientes
    qs = Deudor.objects.filter(activo=True)
    if cartera:
        qs = qs.filter(cartera=cartera)
    if agencia:
        qs = qs.filter(agencia=agencia)
    if distritos:
        qs = qs.filter(distrito__in=distritos)
        
    if estado_negociacion == 'con_negociacion':
        qs = qs.exclude(negociacion__isnull=True).exclude(negociacion__exact='').exclude(negociacion__iexact='nan').exclude(negociacion__iexact='none').exclude(negociacion__iexact='null')
    elif estado_negociacion == 'sin_negociacion':
        qs = qs.filter(Q(negociacion__isnull=True) | Q(negociacion__exact='') | Q(negociacion__iexact='nan') | Q(negociacion__iexact='none') | Q(negociacion__iexact='null'))
        
    # Ordenar por agencia y luego cliente
    clientes = list(qs.order_by('agencia', 'nombre_completo'))
    
    if not clientes:
        return HttpResponse("No se encontraron clientes con esos filtros.", status=404)

    # Agrupar clientes por agencia
    agencias_dict = {}
    for c in clientes:
        ag = c.agencia or 'SIN AGENCIA'
        if ag not in agencias_dict:
            agencias_dict[ag] = []
        agencias_dict[ag].append(c)

    plantilla_path = os.path.join(settings.BASE_DIR, 'REPORTE DE GESTIONES JUDICALES Y EXTRAJUDICIALES AG. SAN BORJA.xlsx')
    
    if not os.path.exists(plantilla_path):
        return HttpResponse("La plantilla Excel no se encontró en el servidor.", status=500)

    # Crear ZIP en memoria
    zip_buffer = io.BytesIO()
    meses_es = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for ag, lista_clientes in agencias_dict.items():
            wb = openpyxl.load_workbook(plantilla_path)
            ws = wb.active
            
            # La cabecera está en la fila 5. Comenzamos a escribir en la 6.
            start_row = 6
            
            # Estilos basicos para celdas de datos
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for idx, c in enumerate(lista_clientes):
                row_idx = start_row + idx
                
                # Datos del cliente
                tiene_convenio = False
                if c.negociacion and str(c.negociacion).lower() not in ('nan', 'none', 'null', ''):
                    tiene_convenio = True
                
                if tiene_convenio and c.ultimo_dia_pago:
                    # Logica dia habil anterior
                    d = c.ultimo_dia_pago - datetime.timedelta(days=1)
                    if d.weekday() == 6: # Sunday
                        d -= datetime.timedelta(days=2)
                    elif d.weekday() == 5: # Saturday
                        d -= datetime.timedelta(days=1)
                    fecha_str = d.strftime('%d/%m/%Y')
                    mes_actual = meses_es[datetime.date.today().month - 1]
                    gestion_txt = f"CON FECHA {fecha_str} SE HACE REQUERIMIENTO DE PAGO PUNTUAL DE SU CUOTA DE CONVENIO DEL MES DE {mes_actual}"
                else:
                    fecha_str = datetime.date.today().strftime('%d/%m/%Y')
                    gestion_txt = f"CON FECHA {fecha_str} SE HACE REQUERIMIENTO DE PAGO DE SU DEUDA EN SU VIVIENDA"

                # Llenar celdas (1-indexed en openpyxl)
                ws.cell(row=row_idx, column=1, value=idx + 1)
                ws.cell(row=row_idx, column=2, value=c.agencia)
                ws.cell(row=row_idx, column=3, value=c.cuenta)
                ws.cell(row=row_idx, column=4, value=c.nombre_completo)
                ws.cell(row=row_idx, column=5, value="") # JUDICIAL
                ws.cell(row=row_idx, column=6, value=gestion_txt)
                
                # Aplicar bordes
                for col_i in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col_i)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Guardar el excel en memoria
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Añadir al zip
            safe_ag_name = "".join([x if x.isalnum() or x in " .-_" else "_" for x in ag])
            zip_file.writestr(f"REPORTE_GESTIONES_{safe_ag_name}.xlsx", excel_buffer.read())

    zip_buffer.seek(0)
    
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="Reportes_Caja_Huancayo_{datetime.date.today().strftime("%Y%m%d")}.zip"'
    return response

"""

filepath = r"C:\CRM PYP\cobranza\views.py"
with open(filepath, 'r', encoding='utf-8') as f:
    content = f.read()

# Make sure not to duplicate
if "def descargar_reporte_excel" not in content:
    with open(filepath, 'a', encoding='utf-8') as f:
        f.write("\n")
        f.write(view_code)
        print("View added to views.py")
else:
    print("View already exists.")
